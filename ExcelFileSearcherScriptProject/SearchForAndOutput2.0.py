# THIS CODE IS ORCHESTRATED BY EMMANUEL OKOSE IN OPTIMUS BANK

import openpyxl

def search_and_output():
    # User input for file paths
    file1 = input("Enter the path to the first Excel file: ")
    file2 = input("Enter the path to the second Excel file: ")

    # User input for column numbers
    col1 = int(input("Enter the column number in the first file to use for searching (1-based): "))
    col2 = int(input("Enter the column to be searched in the second file (1-based): "))
    col3 = int(input("Enter the column to get content from the second file (1-based): "))

    # User input for output file
    output_file = input("Enter the desired name for the output file: ")

    # Open and access Excel files
    wb1 = openpyxl.load_workbook(file1)
    sheet1 = wb1.active
    wb2 = openpyxl.load_workbook(file2)
    sheet2 = wb2.active

    # Create output workbook and sheet
    wb_out = openpyxl.Workbook()
    sheet_out = wb_out.active

    # Write headers
    # sheet_out.cell(row=1, column=1).value = f"Item from {file1} ({col1})"
    sheet_out.cell(row=1, column=1).value = f"Match in {file2} ({col2})"
    sheet_out.cell(row=1, column=2).value = f"Value from {file2} ({col3})"

    row_out = 2

    # Iterate through first file rows
    for row in range(2, sheet1.max_row + 1):
        item = sheet1.cell(row=row, column=col1).value
        found = False

        # Iterate through second file rows to find matches
        for row2 in range(2, sheet2.max_row + 1):
            if sheet2.cell(row=row2, column=col2).value == item:
                found = True
                match_value = sheet2.cell(row=row2, column=col3).value
             #   sheet_out.cell(row=row_out, column=1).value = item
                sheet_out.cell(row=row_out, column=1).value = item
                sheet_out.cell(row=row_out, column=2).value = match_value
                row_out += 1
                break

        # Write output if no match found
        if not found:
         #   sheet_out.cell(row=row_out, column=1).value = item
            sheet_out.cell(row=row_out, column=1).value = item
            sheet_out.cell(row=row_out, column=2).value = "NOT FOUND"
            row_out += 1

    # Save output file
    wb_out.save(output_file)

    print(f"Results written to {output_file}")

# Run the function
print("==========================================")
print("| Made with love by Emmanuel .C. Okose...|")
print("==========================================")
print("")
search_and_output()
